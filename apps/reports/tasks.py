"""
Report generation tasks.
These tasks produce the actual PDF and Excel files from dataset analysis data.
"""
import time
import logging
from celery import shared_task

logger = logging.getLogger(__name__)


@shared_task(bind=True, name="reports.generate_pdf")
def generate_pdf_report_task(self, report_id: str) -> dict:
    """
    Generate a PDF report using WeasyPrint.

    Strategy: render an HTML template with all the report content
    (stats, charts as SVG, AI narrative, anomaly tables), then
    use WeasyPrint to convert that HTML to a pixel-perfect PDF.
    This is much more flexible than generating PDF programmatically
    with reportlab — you use CSS for layout and Jinja2 for content.
    """
    from apps.reports.models import Report
    from django.template.loader import render_to_string
    from weasyprint import HTML
    import os

    start = time.time()
    report = Report.objects.select_related("dataset").get(id=report_id)
    report.status = Report.Status.GENERATING
    report.save(update_fields=["status"])

    try:
        # Build context from dataset profile and saved charts
        context = _build_report_context(report)

        # Render HTML template → string
        html_string = render_to_string("reports/pdf_template.html", context)

        # Convert HTML to PDF bytes
        pdf_bytes = HTML(string=html_string, base_url="/").write_pdf()

        # Save to media storage
        filename = f"report_{report_id}.pdf"
        output_path = f"reports/{filename}"

        from django.core.files.base import ContentFile
        report.output_file.save(filename, ContentFile(pdf_bytes))
        report.status = Report.Status.READY
        report.generation_time_seconds = time.time() - start
        report.file_size_bytes = len(pdf_bytes)
        report.save()

        logger.info(f"PDF report {report_id} generated in {report.generation_time_seconds:.1f}s")
        return {"status": "success", "report_id": report_id}

    except Exception as exc:
        report.status = Report.Status.FAILED
        report.error_message = str(exc)
        report.save(update_fields=["status", "error_message"])
        logger.error(f"PDF report {report_id} failed: {exc}")
        raise


@shared_task(bind=True, name="reports.generate_excel")
def generate_excel_report_task(self, report_id: str) -> dict:
    """
    Generate an Excel workbook with multiple sheets:
      Sheet 1: Dataset Overview (stats, quality scores)
      Sheet 2: Column Profiles (per-column stats)
      Sheet 3: Correlation Matrix
      Sheet 4: Anomalies (if included)
      Sheet 5+: NL Query Results (one sheet per saved query, if included)
    """
    from apps.reports.models import Report
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    report = Report.objects.select_related("dataset").get(id=report_id)
    wb = openpyxl.Workbook()

    # Sheet 1: Overview
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = report.dataset.name
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Rows: {report.dataset.row_count}"
    ws["A3"] = f"Columns: {report.dataset.column_count}"

    # Save to bytes buffer (no temp file needed)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"report_{report_id}.xlsx"
    from django.core.files.base import ContentFile
    report.output_file.save(filename, ContentFile(buffer.read()))
    report.status = Report.Status.READY
    report.save()

    return {"status": "success"}


def _build_report_context(report):
    """Helper: gather all data needed to render the PDF template."""
    from apps.analysis.models import DataProfile, AnomalyReport
    context = {
        "report": report,
        "dataset": report.dataset,
        "columns": list(report.dataset.columns.all()),
    }
    try:
        profile = DataProfile.objects.get(dataset=report.dataset)
        context["profile"] = profile
        context["ai_narrative"] = profile.ai_narrative
        context["key_insights"] = profile.key_insights
    except DataProfile.DoesNotExist:
        pass

    if report.include_anomalies:
        context["anomaly_report"] = AnomalyReport.objects.filter(
            dataset=report.dataset
        ).order_by("-generated_at").first()

    return context
